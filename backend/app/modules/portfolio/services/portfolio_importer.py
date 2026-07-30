"""Transaction and holding file parsers (CDSL CAS, Groww, Zerodha)."""

import io
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from app.core.logging import logger

# Accepted column names (case-insensitive) -> internal key
_COL_MAP = {
    # Generic / common
    "date": "date", "trade date": "date", "trade_date": "date",
    "transaction date": "date", "transaction_date": "date",
    "symbol": "symbol", "ticker": "symbol", "scrip": "symbol", "stock": "symbol",
    "type": "type", "transaction type": "type", "transaction_type": "type",
    "trade type": "type", "trade_type": "type", "action": "type",
    "qty": "quantity", "quantity": "quantity", "shares": "quantity", "units": "quantity",
    "price": "price", "rate": "price", "trade price": "price", "avg price": "price",
    # Zerodha contract note
    "instrument": "symbol", "avg. price": "price", "net price": "price",
    "buy/sell": "type", "series": "_ignore",
    # Zerodha console Tradebook export (CSV/XLSX)
    "trade id": "_order_id",
    # Groww trade statement (CSV)
    "stock symbol": "symbol", "average traded price": "price",
    "stock/scrip name": "_name",
    # Groww stock order history (XLSX)
    "stock name": "_name",
    "isin": "_isin",
    "segment": "_segment",
    "value": "_total",
    "exchange": "_exchange",
    "exchange order id": "_order_id",
    "execution date and time": "date",
    "order status": "_status",
    # Groww MF order history (CSV/XLSX)
    "fund name": "_name", "scheme name": "_name", "scheme": "_name",
    "nav": "price", "nav (rs)": "price", "nav(rs.)": "price",
    "order date": "date", "allotment date": "date",
    "order type": "type",
    "units allotted": "quantity", "units purchased": "quantity",
    "units redeemed": "quantity",
    "amount (rs)": "_total", "amount(rs.)": "_total", "amount": "_total",
    "folio no": "_folio", "folio no.": "_folio", "folio number": "_folio",
    "order id": "_order_id",
    # Binance trade history
    "date(utc)": "date", "pair": "symbol", "side": "type",
    "executed": "quantity", "fee": "_fee",
    # Groww holdings-snapshot exports — recognised only so _detect_broker can
    # flag them as unsupported (see _UNSUPPORTED_GROWW_SHAPES), not parsed as
    # transactions. Mapped as _ignore purely so the XLSX header-row finder's
    # >=3-match threshold locates the real header row under the "HOLDING
    # SUMMARY" preamble section.
    "average buy price": "_ignore", "buy value": "_ignore",
    "closing price": "_ignore", "closing value": "_ignore",
    "unrealised p&l": "_ignore",
    "amc": "_ignore", "category": "_ignore", "sub-category": "_ignore",
    "source": "_ignore", "invested value": "_ignore",
    "current value": "_ignore", "returns": "_ignore", "xirr": "_ignore",
}

# Detected shapes that are real Groww exports but point-in-time holdings
# snapshots (no transaction date/type), not transaction logs — this generic
# per-row transaction parser rejects them with a specific message pointing at
# the dedicated snapshot-import endpoints (parse_groww_stocks_holdings /
# parse_groww_mf_holdings, below), rather than either running them through
# the wrong parser or silently failing.
_UNSUPPORTED_GROWW_SHAPES = {
    "groww_holdings_snapshot": (
        "This looks like a Groww Stocks Holdings Statement — a point-in-time "
        "holdings snapshot, not a transaction log, so it can't be imported "
        "via this endpoint. Use POST /portfolios/{id}/import/groww/holdings "
        "instead, or connect Groww's live sync under Settings if API "
        "credentials are available."
    ),
    "groww_mf_holdings_snapshot": (
        "This looks like a Groww Mutual Funds holdings summary — a "
        "point-in-time snapshot, not a transaction log, so it can't be "
        "imported via this endpoint. Use "
        "POST /portfolios/{id}/import/groww/mf-holdings instead, or connect "
        "Groww's live sync under Settings if API credentials are available."
    ),
}

_VALID_TYPES = {"buy", "sell", "dividend", "interest", "split", "bonus", "contribution", "withdrawal"}

_TYPE_ALIAS = {
    "b": "buy", "purchase": "buy", "bought": "buy",
    "lumpsum": "buy", "additional purchase": "buy", "sip": "buy",
    "switch in": "buy", "switch-in": "buy",
    "s": "sell", "sale": "sell", "sold": "sell",
    "redeem": "sell", "redemption": "sell",
    "switch out": "sell", "switch-out": "sell",
    "d": "dividend", "div": "dividend",
}

def _detect_broker(header: List[str]) -> Optional[str]:
    lowers = {c.strip().lower() for c in header}
    if "pair" in lowers and ("date(utc)" in lowers or "side" in lowers):
        return "binance"
    if "average buy price" in lowers and "stock name" in lowers:
        return "groww_holdings_snapshot"
    if "amc" in lowers and "xirr" in lowers and "scheme name" in lowers:
        return "groww_mf_holdings_snapshot"
    if ("fund name" in lowers or "scheme name" in lowers) and ("nav" in lowers or "nav (rs)" in lowers or "nav(rs.)" in lowers):
        return "groww_mf"
    if "execution date and time" in lowers:
        return "groww"
    if "stock symbol" in lowers or "average traded price" in lowers:
        return "groww"
    if ("instrument" in lowers or "avg. price" in lowers) and "series" in lowers:
        return "zerodha"
    # Zerodha console Tradebook export (Equity or Mutual Fund)
    if "trade type" in lowers and "segment" in lowers and "exchange" in lowers and (
        "trade id" in lowers or "order id" in lowers
    ):
        return "zerodha"
    return None

def _mf_symbol(name: str) -> str:
    slug = re.sub(r"[^A-Z0-9]+", "_", name.upper().strip())
    return slug[:40].rstrip("_") + "_MF"

def _normalise_binance_symbol(pair: str) -> str:
    from app.core.binance import split_quote_asset

    base, quote = split_quote_asset(pair)
    if base is None:
        return pair
    return f"{base}-{quote}"

def _normalise_type(raw: str) -> str:
    v = raw.strip().lower()
    return _TYPE_ALIAS.get(v, v)

def _parse_date(raw: str) -> Optional[datetime]:
    for fmt in (
        "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%Y/%m/%d",
        "%d-%b-%Y", "%d %B %Y",
        "%d-%m-%Y %I:%M %p", "%d-%m-%Y %H:%M",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
    ):
        try:
            # Parse as naive and localize to UTC
            naive = datetime.strptime(raw.strip(), fmt)
            return naive.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None

def _validate_row(row: Dict[str, Any], idx: int) -> List[str]:
    errs = []
    if not row.get("symbol"):
        errs.append(f"row {idx}: missing symbol")
    if not row.get("date"):
        errs.append(f"row {idx}: missing or unparseable date")
    txn_type = row.get("type", "")
    if txn_type not in _VALID_TYPES:
        errs.append(f"row {idx}: invalid type '{txn_type}' — expected one of {sorted(_VALID_TYPES)}")
    try:
        qty = float(row.get("quantity", 0))
        if qty <= 0:
            errs.append(f"row {idx}: quantity must be > 0")
    except (TypeError, ValueError):
        errs.append(f"row {idx}: quantity is not a number")
    try:
        price = float(row.get("price", 0))
        if price < 0:
            errs.append(f"row {idx}: price must be ≥ 0")
    except (TypeError, ValueError):
        errs.append(f"row {idx}: price is not a number")
    return errs

def _rows_from_records(records: List[Dict[str, Any]], broker: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows, errors = [], []

    if not broker and records:
        broker = _detect_broker(list(records[0].keys()))

    if broker in _UNSUPPORTED_GROWW_SHAPES:
        return [], [_UNSUPPORTED_GROWW_SHAPES[broker]]

    if records:
        found_cols = list(records[0].keys())
        recognised = [c for c in found_cols if _COL_MAP.get(c.strip().lower())]
        logger.debug(f"importer columns found={found_cols} recognised={recognised} broker={broker}")
        if not recognised:
            return [], [
                f"No recognised columns found. File headers: {found_cols}."
            ]

    for i, rec in enumerate(records, start=2):
        normalised: Dict[str, Any] = {}
        extras: Dict[str, str] = {}
        for raw_col, val in rec.items():
            key = _COL_MAP.get((raw_col or "").strip().lower())
            if key is None:
                continue
            str_val = str(val).strip() if val is not None else ""
            if key.startswith("_"):
                extras[key] = str_val
            else:
                normalised[key] = str_val

        if not normalised:
            continue
        if not any(v.strip() for v in normalised.values() if isinstance(v, str)):
            continue

        if broker == "groww_mf":
            status = extras.get("_status", "").strip().lower()
            if status and status not in ("executed", "allotted", "redeemed", "completed", "successful", "success"):
                continue
            if not normalised.get("symbol") and extras.get("_name"):
                normalised["symbol"] = _mf_symbol(extras["_name"])

        if broker == "groww" and extras.get("_status", "").strip().lower() != "executed":
            if extras.get("_status"):
                continue

        if broker == "binance" and "quantity" in normalised:
            normalised["quantity"] = normalised["quantity"].split()[0]
        if broker == "binance" and "symbol" in normalised:
            normalised["symbol"] = _normalise_binance_symbol(normalised["symbol"])

        is_mf_segment = extras.get("_segment", "").strip().upper() == "MF"
        if is_mf_segment and normalised.get("symbol"):
            if not extras.get("_name"):
                extras["_name"] = normalised["symbol"]
            normalised["symbol"] = _mf_symbol(normalised["symbol"])

        if (
            broker in ("zerodha", "groww")
            and extras.get("_segment", "").strip().upper() != "MF"
            and normalised.get("symbol")
            and not normalised["symbol"].upper().endswith((".NS", ".BO"))
        ):
            # NSE/BSE holdings of the same stock are fungible in an Indian demat
            # account — always canonicalise to NSE regardless of which exchange
            # a given historical trade executed on, so buys/sells of the same
            # stock across exchanges net into one position instead of splitting
            # into unrelated .NS/.BO symbols.
            normalised["symbol"] = f"{normalised['symbol'].upper()}.NS"

        if "price" not in normalised and "_total" in extras:
            try:
                total = float(extras["_total"].replace(",", ""))
                qty = float(normalised.get("quantity", "0"))
                if qty > 0:
                    normalised["price"] = str(round(total / qty, 4))
            except (ValueError, ZeroDivisionError):
                pass

        if "date" in normalised:
            parsed = _parse_date(normalised["date"])
            normalised["date"] = parsed
        if "type" in normalised:
            normalised["type"] = _normalise_type(normalised["type"])

        errs = _validate_row(normalised, i)
        errors.extend(errs)
        if not errs:
            rows.append({
                "symbol":           normalised["symbol"].upper(),
                "type":             normalised["type"].upper(),  # Store as uppercase
                "quantity":         float(normalised["quantity"]),
                "price":            float(normalised["price"]),
                "date":             normalised["date"],
                "broker":           broker or "import",
                "broker_reference": extras.get("_order_id") or None,
                "isin":             extras.get("_isin") or None,
                "exchange":         extras.get("_exchange") or None,
                "name":             extras.get("_name") or None,
                "asset_type":       "mutual_fund" if (broker == "groww_mf" or is_mf_segment) else None,
            })

    return rows, errors

def _parse_csv(content: bytes, broker: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    import csv
    text = content.decode("utf-8-sig", errors="replace")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    reader = csv.DictReader(io.StringIO(text))
    return _rows_from_records(list(reader), broker=broker)

def _parse_xlsx(content: bytes, broker: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl not installed — cannot parse XLSX files"]

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return [], ["Empty spreadsheet"]

    # Find the header row
    def _find_header_row(rows_list: List[Any]) -> int:
        for i, row in enumerate(rows_list):
            cols = [str(c).strip().lower() for c in row if c is not None]
            if sum(1 for c in cols if c in _COL_MAP) >= 3:
                return i
        return 0

    header_idx = _find_header_row(rows_raw)
    header = [str(c) if c is not None else "" for c in rows_raw[header_idx]]
    records = []
    for row in rows_raw[header_idx + 1:]:
        rec = {header[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(header)}
        records.append(rec)

    return _rows_from_records(records, broker=broker)

def _parse_pdf(content: bytes, broker: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        import pdfplumber
    except ImportError:
        return [], ["pdfplumber not installed — cannot parse PDF files"]

    records = []
    header: List[str] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            if not header:
                header = [str(c or "") for c in table[0]]
            for row in table[1:]:
                if row:
                    rec = {header[i]: str(v or "") for i, v in enumerate(row) if i < len(header)}
                    records.append(rec)

    if not records:
        return [], ["No tables found in PDF. Verify the file contains a transaction table."]

    rows, errors = _rows_from_records(records, broker=broker)
    return rows, errors

def parse_transaction_file(content: bytes, ext: str, broker: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    if ext in ("xlsx", "xls"):
        return _parse_xlsx(content, broker=broker)
    if ext == "pdf":
        return _parse_pdf(content, broker=broker)
    return _parse_csv(content, broker=broker)


# ── CDSL CAS Parser ───────────────────────────────────────────────────────────

def _num(val: Any) -> Optional[float]:
    if val is None:
        return None
    s = str(val).replace(",", "").strip()
    if s in ("--", "-", "", "N/A", "NA"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def _cell(row: List[Any], idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).replace("\n", " ").strip()

def _clean_isin(raw: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", raw.upper())

def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s).lower()).strip("_")

def _clean_scheme_name(raw: str) -> str:
    s = raw.replace("\n", " ").strip()
    if "#" in s:
        after = s.split("#", 1)[1].strip()
        after = re.sub(r"^[A-Z0-9 ]+ MF-", "", after, flags=re.IGNORECASE).strip()
        return after if after else s
    return s

def _is_folio_header(row: List[Any]) -> bool:
    norms = {_norm(c) for c in row if c}
    return (
        any("isin" in n for n in norms) and
        any("folio" in n for n in norms) and
        any("scheme" in n for n in norms) and
        any("nav" in n for n in norms)
    )

def _is_holding_header(row: List[Any]) -> bool:
    norms = {_norm(c) for c in row if c}
    return (
        any("isin" in n for n in norms) and
        any("security" in n for n in norms) and
        any("currentbal" in n or (n.startswith("current") and "bal" in n) for n in norms) and
        any("marketprice" in n or "faceval" in n or ("market" in n and "price" in n) for n in norms)
    )

def _map_folio_cols(header_row: List[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for ci, cell in enumerate(header_row):
        n = _norm(cell)
        raw = str(cell).lower()
        if "scheme" in n and "name" in n:
            m.setdefault("scheme", ci)
        elif n == "isin" or n == "isinisin":
            m.setdefault("isin", ci)
        elif "folio" in n:
            m.setdefault("folio", ci)
        elif "closing" in n or ("unit" in n and "closing" in n):
            m.setdefault("units", ci)
        elif "nav" in n and "cumul" not in n and "unreali" not in n and "valuation" not in n:
            m.setdefault("nav", ci)
        elif "cumul" in n or "invest" in n:
            m.setdefault("invested", ci)
        elif "valuation" in n:
            m.setdefault("valuation", ci)
        elif "unreali" in n and "%" not in raw:
            m.setdefault("pnl", ci)
        elif "unreali" in n and "%" in raw:
            m.setdefault("pnl_pct", ci)
    return m

def _map_holding_cols(header_row: List[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for ci, cell in enumerate(header_row):
        n = _norm(cell)
        if n in ("isin", "isinisin"):
            m.setdefault("isin", ci)
        elif "security" in n:
            m.setdefault("security", ci)
        elif "currentbal" in n or (n.startswith("current") and len(n) < 20):
            m.setdefault("units", ci)
        elif "marketprice" in n or "faceval" in n or ("market" in n and "price" in n):
            m.setdefault("price", ci)
        elif "value" in n and "pledge" not in n and "setup" not in n and "face" not in n:
            m.setdefault("value", ci)
    return m

def _parse_folio_table(table: List[List[Any]]) -> List[Dict[str, Any]]:
    hdr_row, data_start = None, None
    for ri in range(min(3, len(table))):
        if _is_folio_header(table[ri]):
            hdr_row = table[ri]
            data_start = ri + 1
            break
    if hdr_row is None:
        return []
    col = _map_folio_cols(hdr_row)
    g = col.get
    results = []
    for row in table[data_start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        scheme = _cell(row, g("scheme", 0))
        if not scheme or "grand total" in scheme.lower() or "load structure" in scheme.lower():
            continue
        isin    = _clean_isin(_cell(row, g("isin", 1)))
        folio   = _cell(row, g("folio",    2))
        units   = _num(_cell(row, g("units",    3)))
        nav     = _num(_cell(row, g("nav",      4)))
        invested= _num(_cell(row, g("invested", 5)))
        val     = _num(_cell(row, g("valuation",6)))
        _num(_cell(row, g("pnl",      7)))
        _num(_cell(row, g("pnl_pct",  8)))

        if units is None or units <= 0:
            continue

        avg_nav = round(invested / units, 4) if units > 0 and invested and invested > 0 else 0.0
        results.append({
            "scheme_name": scheme,
            "isin": isin,
            "folio_no": folio,
            "units": units,
            "avg_nav": avg_nav,
            "current_nav": nav or 0.0,
            "valuation": val or 0.0,
            "source": "cas_folio"
        })
    return results

def _parse_holding_table(table: List[List[Any]], dp_name: str) -> List[Dict[str, Any]]:
    hdr_row, data_start = None, None
    for ri in range(min(3, len(table))):
        if _is_holding_header(table[ri]):
            hdr_row = table[ri]
            data_start = ri + 1
            break
    if hdr_row is None:
        return []
    col = _map_holding_cols(hdr_row)
    g = col.get
    results = []
    for row in table[data_start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        isin = _clean_isin(_cell(row, g("isin", 0)))
        if not isin.startswith("INF"):
            continue
        security_raw = _cell(row, g("security", 1))
        units = _num(_cell(row, g("units", 2)))
        price = _num(_cell(row, g("price", 7)))
        value = _num(_cell(row, g("value", 8)))

        if units is None or units <= 0:
            continue
        results.append({
            "scheme_name": _clean_scheme_name(security_raw),
            "isin": isin,
            "folio_no": "",
            "units": units,
            "avg_nav": 0.0,
            "current_nav": price or 0.0,
            "valuation": value or 0.0,
            "source": "cas_demat",
            "dp": dp_name
        })
    return results

def parse_cdsl_cas(content: bytes, password: Optional[str] = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber not installed — cannot parse CDSL CAS PDF")

    open_kwargs = {}
    if password:
        open_kwargs["password"] = password

    from pdfminer.pdfdocument import PDFPasswordIncorrect

    try:
        pdf = pdfplumber.open(io.BytesIO(content), **open_kwargs)
    except Exception as exc:
        cause = exc.args[0] if exc.args else None
        if isinstance(exc, PDFPasswordIncorrect) or isinstance(cause, PDFPasswordIncorrect):
            marker = "PDF_PASSWORD_INCORRECT" if password else "PDF_PASSWORD_REQUIRED"
            raise ValueError(marker) from exc
        raise ValueError(f"Cannot open PDF: {exc}") from exc

    mf_folios = []
    demat_mf = []
    current_dp = ""

    _dp_pattern = re.compile(
        r"DP\s+Name\s*[:\-]\s*([A-Z][A-Z0-9 &()]+?)(?:\s{3,}|BO ID|$)",
        re.IGNORECASE,
    )

    with pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            tables = page.extract_tables() or []

            dm = _dp_pattern.search(text)
            if dm:
                cand = dm.group(1).strip().split("\n")[0].strip()
                if len(cand) > 3:
                    current_dp = cand

            for table in tables:
                if not table:
                    continue
                if any(_is_folio_header(table[ri]) for ri in range(min(3, len(table)))):
                    mf_folios.extend(_parse_folio_table(table))
                elif any(_is_holding_header(table[ri]) for ri in range(min(3, len(table)))):
                    demat_mf.extend(_parse_holding_table(table, current_dp))

    # Deduplicate demat
    seen = set()
    deduped_demat = []
    for h in demat_mf:
        key = (h["isin"], h["dp"])
        if key not in seen:
            seen.add(key)
            deduped_demat.append(h)

    # Merge by ISIN
    merged = {}
    for h in mf_folios:
        key = h["isin"] or h["folio_no"]
        merged[key] = {
            "isin": h["isin"],
            "scheme_name": h["scheme_name"],
            "units": h["units"],
            "avg_nav": h["avg_nav"],
            "current_nav": h["current_nav"],
            "source": "CDSL CAS (Folio)",
        }

    for h in deduped_demat:
        key = h["isin"]
        if key in merged:
            merged[key]["units"] += h["units"]
        else:
            merged[key] = {
                "isin": h["isin"],
                "scheme_name": h["scheme_name"],
                "units": h["units"],
                "avg_nav": 0.0,
                "current_nav": h["current_nav"],
                "source": "CDSL CAS (Demat)",
            }

    payloads = []
    for key, m in merged.items():
        isin = m["isin"]
        units = m["units"]
        avg_nav = m["avg_nav"]
        current_nav = m["current_nav"]
        symbol = f"{isin}_MF" if isin else f"{_slug(m['scheme_name'])}_MF"
        payloads.append({
            "symbol": symbol,
            "name": m["scheme_name"],
            "quantity": units,
            "avg_buy_price": avg_nav,
            "current_price": current_nav if current_nav else None,
            "source": m["source"],
            "asset_type": "mutual_fund",
        })

    summary = {
        "mf_folios_count": len(mf_folios),
        "demat_mf_count": len(deduped_demat),
        "merged_count": len(payloads),
    }

    return payloads, summary


# ── Groww Holdings-Snapshot Parsers ───────────────────────────────────────────
# Both are point-in-time snapshot exports (no transaction date/type), unlike
# the Stock/MF Order History shapes handled by parse_transaction_file above —
# same broker_snapshot modelling as CDSL CAS/NPS/EPF, not the per-row ledger
# path. Detection uses _norm() (strips all non-alnum) rather than the generic
# importer's exact-string _COL_MAP, so "Folio No." (trailing period) matches
# "folio no" without a separate mapping entry — the header-detection issue
# found during scoping doesn't reoccur here by construction.

def _find_header_row(rows_raw: List[Any], required_norms: List[str]) -> Optional[int]:
    for i, row in enumerate(rows_raw):
        norms = {_norm(c) for c in row if c is not None}
        if all(req in norms for req in required_norms):
            return i
    return None

def parse_groww_stocks_holdings(content: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Groww "Stocks Holdings Statement" export: Stock Name, ISIN, Quantity,
    Average buy price, Buy value, Closing price, Closing value, Unrealised P&L.
    No ticker/exchange column — unlike Stock Order History, this format only
    ever gives a company name + ISIN. There is no ISIN->ticker resolution
    anywhere in this codebase (same confirmed gap as Zerodha's Tax P&L/Holdings
    Statement exports — see PROVIDERS.md Known Backlog #2), so the symbol is
    synthesised from ISIN (falling back to a name slug), matching the same
    ISIN-preferred/name-slug-fallback pattern parse_cdsl_cas uses for MF
    holdings. This intentionally does NOT resolve to the real NSE/BSE ticker
    a live sync or Order History import would use for the same stock — same
    "synthetic symbol, distinct namespace" convention already used by NPS/EPF
    for sources with no live-market ticker of their own. A stock imported both
    ways will show up as two separate positions until a real ISIN->ticker
    mapping exists; flagged, not silently merged."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl not installed — cannot parse XLSX files")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))

    header_idx = _find_header_row(rows_raw, ["stockname", "isin", "quantity"])
    if header_idx is None:
        raise ValueError(
            "Could not find the holdings table (expected Stock Name/ISIN/Quantity "
            "columns) in this file — is this a Groww Stocks Holdings Statement export?"
        )

    header = rows_raw[header_idx]
    col: Dict[str, int] = {}
    for ci, cell in enumerate(header):
        n = _norm(cell)
        if n == "stockname":
            col.setdefault("name", ci)
        elif n == "isin":
            col.setdefault("isin", ci)
        elif n == "quantity":
            col.setdefault("quantity", ci)
        elif n == "averagebuyprice":
            col.setdefault("avg_price", ci)
        elif n == "closingprice":
            col.setdefault("current_price", ci)
    g = col.get

    payloads = []
    for row in rows_raw[header_idx + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        name = _cell(row, g("name", 0))
        if not name:
            continue
        isin = _clean_isin(_cell(row, g("isin", 1)))
        qty = _num(_cell(row, g("quantity", 2)))
        avg_price = _num(_cell(row, g("avg_price", 3)))
        current_price = _num(_cell(row, g("current_price", -1))) if "current_price" in col else None

        if qty is None or qty <= 0:
            continue

        symbol = f"{isin}_HOLDING" if isin else f"{_slug(name).upper()}_HOLDING"
        payloads.append({
            "symbol": symbol,
            "name": name,
            "quantity": qty,
            "avg_buy_price": avg_price or 0.0,
            "current_price": current_price,
            "asset_type": "equity",
        })

    return payloads, {"rows_found": len(payloads)}

def parse_groww_mf_holdings(content: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Groww Mutual Funds holdings summary: Scheme Name, AMC, Category,
    Sub-category, Folio No., Source, Units, Invested Value, Current Value,
    Returns, XIRR — preceded by a "HOLDING SUMMARY" totals section that must
    not be mistaken for the real header (it only shares an XIRR column).
    No per-unit NAV column here (unlike CDSL's MF folio table) — avg/current
    NAV are derived from Invested/Current Value ÷ Units. No ISIN column
    either, so symbol reuses the same _mf_symbol() slug convention the
    working MF Order History import already uses — a fund appearing in both
    exports resolves to the same symbol/Asset/Position, unlike the equity
    holdings case above."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl not installed — cannot parse XLSX files")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))

    header_idx = _find_header_row(rows_raw, ["schemename", "units", "foliono"])
    if header_idx is None:
        raise ValueError(
            "Could not find the MF holdings table (expected Scheme Name/Units/"
            "Folio No. columns) in this file — is this a Groww Mutual Funds "
            "holdings summary export?"
        )

    header = rows_raw[header_idx]
    col: Dict[str, int] = {}
    for ci, cell in enumerate(header):
        n = _norm(cell)
        if n == "schemename":
            col.setdefault("name", ci)
        elif n == "foliono":
            col.setdefault("folio", ci)
        elif n == "units":
            col.setdefault("units", ci)
        elif n == "investedvalue":
            col.setdefault("invested", ci)
        elif n == "currentvalue":
            col.setdefault("current", ci)
    g = col.get

    payloads = []
    for row in rows_raw[header_idx + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        name = _cell(row, g("name", 0))
        if not name:
            continue
        folio = _cell(row, g("folio", 4))
        units = _num(_cell(row, g("units", 6)))
        invested = _num(_cell(row, g("invested", 7)))
        current = _num(_cell(row, g("current", 8)))

        if units is None or units <= 0:
            continue

        avg_nav = round(invested / units, 4) if invested and units > 0 else 0.0
        current_nav = round(current / units, 4) if current and units > 0 else None

        payloads.append({
            "symbol": _mf_symbol(name),
            "name": name,
            "quantity": units,
            "avg_buy_price": avg_nav,
            "current_price": current_nav,
            "asset_type": "mutual_fund",
            "folio_no": folio,
        })

    return payloads, {"rows_found": len(payloads)}


# ── NPS Statement Parser ──────────────────────────────────────────────────────

def _num_paren(val: Any) -> Optional[float]:
    """Like _num(), but treats parenthesised amounts as negative, e.g. "(25.55)" -> -25.55."""
    if val is None:
        return None
    s = str(val).replace(",", "").strip()
    if s in ("--", "-", "", "N/A", "NA"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1].strip()
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n

def _is_blank_row(row: List[str]) -> bool:
    return not row or all(not str(c).strip() for c in row)

def _detect_nps_tier(first_line: str) -> Optional[int]:
    m = re.search(r"Tier\s+(I|II)\s+Account", first_line, re.IGNORECASE)
    if not m:
        return None
    return 2 if m.group(1).upper() == "II" else 1

def _nps_scheme_letter(scheme_name: str) -> Optional[str]:
    m = re.search(r"SCHEME\s+([A-Z])\s*-\s*TIER", scheme_name, re.IGNORECASE)
    return m.group(1).upper() if m else None

_NPS_BUY_DESCRIPTIONS = {"by voluntary contributions", "tier-2 contribution"}
_NPS_SKIP_DESCRIPTIONS = {"opening balance", "closing balance"}

def _nps_rows_from_text(text: str) -> List[List[str]]:
    import csv
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return list(csv.reader(io.StringIO(text)))


def _nps_rows_from_xlsx(content: bytes) -> List[List[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    return [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]


def _nps_rows_from_pdf(content: bytes) -> List[List[str]]:
    import pdfplumber
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    return _nps_rows_from_text(text)


def parse_nps_statement(content: bytes, ext: str = "csv") -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Parses an NPS (National Pension System) transaction statement — CSV, XLSX, or PDF.

    The statement is the same comma-structured tabular report regardless of
    export format (a PDF export is a printed rendition of the same
    comma-separated rows), so every format is normalised to the same
    List[List[str]] row grid before the shared section-parsing logic below.

    Returns (holdings, transactions, summary):
    - holdings: current scheme-wise snapshot from "Investment Details - Scheme
      Wise Summary" (one per scheme: symbol, name, tier, quantity, current_nav, as_of_date).
    - transactions: real BUY/SELL rows from "Transaction Details" per-scheme
      sections (Opening balance / Closing Balance rows excluded).
    """
    ext = (ext or "csv").lower().lstrip(".")
    if ext in ("xlsx", "xls"):
        try:
            rows = _nps_rows_from_xlsx(content)
        except ImportError:
            raise ImportError("openpyxl not installed — cannot parse XLSX NPS statement")
    elif ext == "pdf":
        try:
            rows = _nps_rows_from_pdf(content)
        except ImportError:
            raise ImportError("pdfplumber not installed — cannot parse PDF NPS statement")
    else:
        rows = _nps_rows_from_text(content.decode("utf-8-sig", errors="replace"))

    if not rows:
        raise ValueError("Empty NPS statement file")

    tier = None
    for row in rows[:20]:
        if row:
            tier = _detect_nps_tier(row[0])
            if tier:
                break
    if tier is None:
        raise ValueError("Could not detect Tier I/II from statement header")

    pran = None
    for row in rows:
        if row and row[0].strip().lower() == "pran" and len(row) > 1:
            pran = row[1].strip().lstrip("'").strip()
            break
    if not pran:
        raise ValueError("Could not find PRAN in statement")

    # ── Investment Details - Scheme Wise Summary ──
    holdings: List[Dict[str, Any]] = []
    summary_idx = next(
        (i for i, r in enumerate(rows) if r and r[0].strip().lower() == "investment details - scheme wise summary"),
        None,
    )
    if summary_idx is not None:
        header_row = rows[summary_idx + 1]
        as_of_date = None
        for cell in header_row:
            dm = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", cell)
            if dm:
                as_of_date = _parse_date(dm.group(1))
                break

        i = summary_idx + 2
        while i < len(rows) and not _is_blank_row(rows[i]):
            row = rows[i]
            scheme_name = row[0].strip()
            units = _num_paren(row[2]) if len(row) > 2 else None
            nav = _num_paren(row[3]) if len(row) > 3 else None
            letter = _nps_scheme_letter(scheme_name)
            if letter and units is not None:
                symbol = f"NPS-{pran}-{letter}-T{tier}"
                holdings.append({
                    "symbol": symbol,
                    "name": scheme_name,
                    "tier": tier,
                    "quantity": units,
                    "current_nav": nav or 0.0,
                    "as_of_date": as_of_date,
                })
            i += 1

    # ── Transaction Details (per-scheme sections) ──
    transactions: List[Dict[str, Any]] = []
    txn_idx = next(
        (i for i, r in enumerate(rows) if r and r[0].strip().lower() == "transaction details"),
        None,
    )
    if txn_idx is not None:
        i = txn_idx + 1
        current_scheme = None
        while i < len(rows):
            row = rows[i]
            if _is_blank_row(row):
                current_scheme = None
                i += 1
                continue
            if current_scheme is None:
                current_scheme = row[0].strip()
                i += 1
                continue
            if row[0].strip().lower() == "date":
                # header row for this scheme's transaction table
                i += 1
                continue

            date_str = row[0].strip() if len(row) > 0 else ""
            desc = row[1].strip() if len(row) > 1 else ""
            amount = _num_paren(row[2]) if len(row) > 2 else None
            nav = _num_paren(row[3]) if len(row) > 3 else None
            units = _num_paren(row[4]) if len(row) > 4 else None

            desc_lower = desc.lower()
            if desc_lower in _NPS_SKIP_DESCRIPTIONS:
                i += 1
                continue

            date = _parse_date(date_str)
            letter = _nps_scheme_letter(current_scheme)
            if not date or not letter or units is None:
                i += 1
                continue

            symbol = f"NPS-{pran}-{letter}-T{tier}"

            if desc_lower in _NPS_BUY_DESCRIPTIONS:
                txn_type = "BUY"
                quantity = units
                txn_amount = amount
            elif desc_lower.startswith("billing for q"):
                txn_type = "SELL"
                quantity = abs(units)
                txn_amount = abs(amount) if amount is not None else None
            else:
                # Unrecognised description: fall back to sign of units.
                txn_type = "BUY" if units >= 0 else "SELL"
                quantity = abs(units)
                txn_amount = abs(amount) if amount is not None else None
                logger.warning(f"nps importer: unrecognised transaction description '{desc}' — inferred {txn_type}")

            transactions.append({
                "symbol": symbol,
                "type": txn_type,
                "quantity": quantity,
                "price": nav or 0.0,
                "amount": txn_amount,
                "date": date,
                "description": desc,
                "broker_reference": f"{symbol}|{date.date().isoformat()}|{desc}",
            })
            i += 1

    summary = {
        "tier": tier,
        "pran": pran,
        "schemes_count": len(holdings),
        "transactions_parsed": len(transactions),
    }

    return holdings, transactions, summary


# ── EPF Statement Parser ──────────────────────────────────────────────────────

_EPF_ESTAB_RE = re.compile(r"Establishment\s+ID/Name\s*[:|]\s*(\S+)\s*/\s*(.+)")
_EPF_MEMBER_RE = re.compile(r"Member\s+ID/Name\s*[:|]\s*(\S+)\s*/\s*(.+)")
_EPF_UAN_RE = re.compile(r"\bUAN\s*[:|]?\s*(\d+)")
_EPF_FY_RE = re.compile(r"Financial\s+Year\s*-\s*(\d{4}-\d{4})")
_EPF_DATE_RE = re.compile(r"(\d{2}/\d{2}/\d{4})")


def _epf_header_fields(text: str) -> Dict[str, Optional[str]]:
    estab = _EPF_ESTAB_RE.search(text)
    member = _EPF_MEMBER_RE.search(text)
    uan = _EPF_UAN_RE.search(text)
    fy = _EPF_FY_RE.search(text)
    return {
        "establishment_name": estab.group(2).strip() if estab else None,
        "member_name": member.group(2).strip() if member else None,
        "uan": uan.group(1).strip() if uan else None,
        "financial_year": fy.group(1).strip() if fy else None,
    }


def _epf_row_date(label: str) -> Optional[datetime]:
    m = _EPF_DATE_RE.search(label)
    return _parse_date(m.group(1).replace("/", "-")) if m else None


def parse_epf_statement(content: bytes, password: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Parses an EPFO Member Passbook PDF (Page 1: Member Passbook).

    Returns (holdings, transactions, summary):
    - holdings: one per UAN — {symbol: "EPF-{uan}", name, uan, member_name,
      establishment_name, quantity=1.0, current_value, as_of_date}. EPF has no
      per-unit NAV (unlike NPS schemes), so it's modelled as a single lump-sum
      balance — quantity is always 1.0 and current_value is the Employee +
      Employer + Pension closing balance, same convention as a manually-valued
      asset_class (see create_manual_asset).
    - transactions: one per wage-month contribution row, Employee + Employer +
      Pension amounts summed into a single combined amount (mirrors NPS's
      single-amount-per-row transaction shape — NPS has no employee/employer
      split to begin with, so there's no existing precedent to split EPF's
      three-way contribution into separate rows). The breakdown is preserved
      in the transaction's description text. Callers must import these as
      kind="broker_trade", not kind="trade": EPF contributions have no
      quantity concept to accumulate the way recalculate_position's BUY
      replay assumes for genuine unit-based holdings (NPS/CDSL) — the
      holdings snapshot above (kind="broker_snapshot") is what drives
      Position.quantity/value, via recalculate_position's broker_snapshot
      fallback.

    The literal "No Transactions available for the this year" marker (present
    verbatim in real EPFO exports for years with zero activity) is a valid,
    non-error state — `summary["zero_transaction_year"]` is set and
    `transactions` is empty, but `holdings` is still populated from the
    Closing Balance row so the Asset/Position gets created regardless.

    NOTE: only the header + zero-transaction path has been validated against a
    real EPFO passbook export. The populated-transaction-row parsing below
    follows the documented column layout (Wage Month, Date, Type, Particulars,
    EPF Wages, EPS Wages, Employee, Employer, Pension) but has not been
    exercised against a real file with actual contribution rows.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber not installed — cannot parse EPF passbook PDF")

    open_kwargs = {}
    if password:
        open_kwargs["password"] = password

    from pdfminer.pdfdocument import PDFPasswordIncorrect

    try:
        pdf = pdfplumber.open(io.BytesIO(content), **open_kwargs)
    except Exception as exc:
        cause = exc.args[0] if exc.args else None
        if isinstance(exc, PDFPasswordIncorrect) or isinstance(cause, PDFPasswordIncorrect):
            marker = "PDF_PASSWORD_INCORRECT" if password else "PDF_PASSWORD_REQUIRED"
            raise ValueError(marker) from exc
        raise ValueError(f"Cannot open PDF: {exc}") from exc

    opening: Optional[Dict[str, Any]] = None
    closing: Optional[Dict[str, Any]] = None
    txn_rows: List[Dict[str, Any]] = []
    zero_txn_year = False
    page2_total_contribution = 0.0

    with pdf:
        pages = pdf.pages
        if not pages:
            raise ValueError("Empty EPF passbook PDF")

        page1_text = pages[0].extract_text() or ""
        header = _epf_header_fields(page1_text)
        if not header.get("uan"):
            raise ValueError("Could not find UAN in EPF passbook")

        for table in pages[0].extract_tables() or []:
            for row in table:
                if not row or all(c is None or not str(c).strip() for c in row):
                    continue
                label = _cell(row, 0)
                label_lower = label.lower()

                if "no transactions available" in label_lower:
                    zero_txn_year = True
                    continue
                if label_lower.startswith("ob int"):
                    opening = {
                        "date": _epf_row_date(label),
                        "employee": _num(_cell(row, 1)) or 0.0,
                        "employer": _num(_cell(row, 2)) or 0.0,
                        "pension": _num(_cell(row, 3)) or 0.0,
                    }
                    continue
                if label_lower.startswith("closing balance as on"):
                    closing = {
                        "date": _epf_row_date(label),
                        "employee": _num(_cell(row, 1)) or 0.0,
                        "employer": _num(_cell(row, 2)) or 0.0,
                        "pension": _num(_cell(row, 3)) or 0.0,
                    }
                    continue
                if (
                    label_lower.startswith("total contributions")
                    or label_lower.startswith("total transfer")
                    or label_lower.startswith("total withdrawals")
                    or label_lower.startswith("interest details")
                ):
                    continue
                if label_lower in ("wage month", "date", ""):
                    continue

                # Candidate transaction row: Wage Month, Date, Type, Particulars,
                # EPF Wages, EPS Wages, Employee, Employer, Pension.
                date_val = _epf_row_date(_cell(row, 1))
                if not date_val:
                    continue
                employee_amt = _num(_cell(row, 6)) or 0.0
                employer_amt = _num(_cell(row, 7)) or 0.0
                pension_amt = _num(_cell(row, 8)) or 0.0
                combined = employee_amt + employer_amt + pension_amt
                if combined <= 0:
                    continue
                txn_rows.append({
                    "wage_month": label,
                    "date": date_val,
                    "employee": employee_amt,
                    "employer": employer_amt,
                    "pension": pension_amt,
                    "amount": combined,
                })

        # Page 2 (Taxable Data): cross-check only — never imported as transactions.
        if len(pages) > 1:
            for table in pages[1].extract_tables() or []:
                for row in table:
                    if not row:
                        continue
                    label = _cell(row, 0).lower()
                    if not label or label in ("cont. month", "total"):
                        continue
                    contrib = _num(_cell(row, 1))
                    if contrib:
                        page2_total_contribution += contrib

    if closing is None:
        raise ValueError("Could not find Closing Balance in EPF passbook")

    uan = header["uan"]
    symbol = f"EPF-{uan}"
    member_name = header.get("member_name") or uan
    total_balance = closing["employee"] + closing["employer"] + closing["pension"]

    holdings = [{
        "symbol": symbol,
        "name": f"EPF - {member_name}",
        "uan": uan,
        "member_name": member_name,
        "establishment_name": header.get("establishment_name"),
        "quantity": 1.0,
        "current_value": total_balance,
        "as_of_date": closing["date"],
    }]

    transactions = []
    for row in txn_rows:
        desc = f"{row['wage_month']}: Employee ₹{row['employee']:,.0f} | Employer ₹{row['employer']:,.0f}"
        if row["pension"]:
            desc += f" | Pension ₹{row['pension']:,.0f}"
        transactions.append({
            "symbol": symbol,
            "type": "BUY",
            "amount": row["amount"],
            "date": row["date"],
            "description": desc,
            "broker_reference": f"{symbol}|{row['wage_month']}|{row['date'].date().isoformat()}",
        })

    page1_total_contribution = sum(r["employee"] + r["employer"] for r in txn_rows)
    cross_check_ok = zero_txn_year or abs(page1_total_contribution - page2_total_contribution) < 1.0
    if not cross_check_ok:
        logger.warning(
            f"epf importer: page1/page2 contribution total mismatch for UAN={uan} "
            f"(page1={page1_total_contribution}, page2={page2_total_contribution})"
        )

    summary = {
        "uan": uan,
        "member_name": member_name,
        "establishment_name": header.get("establishment_name"),
        "financial_year": header.get("financial_year"),
        "zero_transaction_year": zero_txn_year,
        "transactions_parsed": len(transactions),
        "closing_balance": total_balance,
        "page2_cross_check_ok": cross_check_ok,
    }

    return holdings, transactions, summary
